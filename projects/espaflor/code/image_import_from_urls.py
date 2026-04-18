"""
Image Import from URLs → Odoo CSV

Purpose:
    Downloads images from URLs (Google Drive, etc.), resizes to 1200x1200 JPEG
    quality 82, converts to base64, writes to CSV format for Odoo import.

Input:
    img.xlsx — Excel file with columns:
        - 'Image' (URL or Holded image URL)
        - 'External ID' (target product.template external ID)
        - 'Name' (for logging only)

Output:
    odoo_images_import.csv — CSV ready for Odoo import:
        External ID,Name,Image

Then use split_big_csv.py to break into batches of 100 for actual import
(otherwise Odoo throws "Content too large").

Dependencies:
    pip install openpyxl requests pillow

Usage:
    python image_import_from_urls.py
"""

import base64
import csv
import io
import re
from typing import Optional

import openpyxl
import requests
from PIL import Image

INPUT_XLSX = "img.xlsx"
OUTPUT_CSV = "odoo_images_import.csv"
TARGET_SIZE = (1200, 1200)
JPEG_QUALITY = 82
MAX_BYTES = 5 * 1024 * 1024  # 5 MB safety cap

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; OdooImage/1.0)"
}


def extract_gdrive_id(url: str) -> Optional[str]:
    """Extract Google Drive file ID from various URL formats."""
    patterns = [
        r"[?&]id=([a-zA-Z0-9_-]+)",
        r"/d/([a-zA-Z0-9_-]+)",
        r"/file/d/([a-zA-Z0-9_-]+)",
    ]
    for pat in patterns:
        m = re.search(pat, url)
        if m:
            return m.group(1)
    return None


def normalize_url(url: str) -> str:
    """Convert Google Drive share URLs to direct download."""
    gid = extract_gdrive_id(url)
    if gid:
        return f"https://drive.usercontent.google.com/download?id={gid}&export=view"
    return url


def download_and_process(url: str) -> Optional[str]:
    """Download image, resize, return base64 string or None on error."""
    try:
        url = normalize_url(url)
        resp = requests.get(url, headers=HEADERS, timeout=30, stream=True)
        resp.raise_for_status()

        content = resp.content
        if len(content) > MAX_BYTES:
            print(f"  [skip] too large: {len(content)} bytes")
            return None

        img = Image.open(io.BytesIO(content))
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")

        img.thumbnail(TARGET_SIZE, Image.Resampling.LANCZOS)

        out = io.BytesIO()
        img.save(out, format="JPEG", quality=JPEG_QUALITY, optimize=True)
        b64 = base64.b64encode(out.getvalue()).decode("ascii")
        return b64

    except Exception as e:
        print(f"  [error] {e}")
        return None


def main():
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb.active

    # Find column indices (header row = 1)
    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
    url_col = headers.get("Image")
    ext_id_col = headers.get("External ID")
    name_col = headers.get("Name", 0)

    if url_col is None or ext_id_col is None:
        raise ValueError("Required columns 'Image' and 'External ID' not found")

    ok_count = 0
    err_count = 0

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["External ID", "Name", "Image"])

        for row in ws.iter_rows(min_row=2, values_only=True):
            url = row[url_col]
            ext_id = row[ext_id_col]
            name = row[name_col] if name_col is not None else ""

            if not url or not ext_id:
                continue

            print(f"Processing {ext_id}: {name}")
            b64 = download_and_process(str(url))

            if b64:
                writer.writerow([ext_id, name or "", b64])
                ok_count += 1
            else:
                err_count += 1

    print(f"\nDone. OK: {ok_count}, Errors: {err_count}")
    print(f"Output: {OUTPUT_CSV}")
    print("Next step: run split_big_csv.py to create batches for Odoo import.")


if __name__ == "__main__":
    main()
